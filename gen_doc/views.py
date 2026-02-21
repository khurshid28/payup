import os,  time
import pythoncom
import win32com.client
from datetime import datetime
from io import BytesIO
import qrcode
from django.core.files.base import ContentFile, File
from docx.shared import Mm
from docxtpl import DocxTemplate, InlineImage
from docx2pdf import convert
from openpyxl.reader.excel import load_workbook

from payup import settings
from payup.settings import BASE_DIR, MEDIA_ROOT, GLOBAL_IP
from openpyxl.drawing.image import Image
from django.contrib.auth.decorators import login_required

class GenDocument:
    def __init__(self, generated_document=None, filename=None, shartnoma=None, buyruq=None, dalolatnoma=None,
                 grafik=None, bayonnoma=None, xulosa=None, context=None, application=None):
        self.generated_document = generated_document
        self.filename = filename
        self.shartnoma = shartnoma  # docx shaklidagi shartnoma template
        self.buyruq = buyruq  # docx shaklidagi buyruq template
        self.dalolatnoma = dalolatnoma  # docx shaklidagi dalolatnoma template
        self.grafik = grafik  # docx shaklidagi grafik template
        self.bayonnoma = bayonnoma  # docx shaklidagi grafik template
        # self.xulosa = xulosa  # docx shaklidagi grafik template
        self.context = context  # application meta json
        self.application = application

    def display_info(self):
        print(self.generated_document)
        print(self.filename)
        print("shartnoma_template=", self.shartnoma)
        print(self.buyruq)
        print(self.dalolatnoma)
        print(self.grafik)
        print(self.context)
        print(self.application)
        return "ok"

    # Shartnoma yaratish
    def gen_shartnoma(self):
        # Application ID saqlab olish
        self.generated_document.application_id = self.application.id
        self.generated_document.save()

        # 1. Shablonni ochish
        docx = DocxTemplate(self.shartnoma)

        # 2. Shablonni to'ldirish va saqlash
        self.context['qr_code'] = "{{qr_code}}"
        docx.render(self.context)

        # 3. Faylni xotirada saqlash
        buffer = BytesIO()
        docx.save(buffer)
        buffer.seek(0)  # Faylni boshidan o‘qish uchun

        # 4. Modelga saqlash
        self.generated_document.docx_shartnoma.save(f"{self.filename}_shartnoma.docx",
                                                    ContentFile(buffer.getvalue()), save=True)

        # DOCXga QR-CODE joylashtirish va pdfga convert qilish
        pdf_filename = f"{self.filename}_shartnoma"
        pdf_url = self.generate_qr_pdf(self.generated_document.docx_shartnoma, pdf_filename)
        # Fayl yo‘lini nisbiy qilish (MEDIA_ROOT ni olib tashlash)
        relative_path = os.path.relpath(pdf_url, settings.MEDIA_ROOT).replace("\\", "/")
        # Faylni modelga saqlash
        with open(pdf_url, "rb") as f:
            self.generated_document.pdf_shartnoma.save(relative_path, File(f), save=True)

        return self.generated_document

    # Buyruq yaratish
    def gen_buyruq(self):
        # 1. Shablonni ochish
        docx = DocxTemplate(self.buyruq)

        # 2. Shablonni to'ldirish va saqlash
        self.context['qr_code'] = "{{qr_code}}"
        docx.render(self.context)

        # 3. Faylni xotirada saqlash
        buffer = BytesIO()
        docx.save(buffer)
        buffer.seek(0)  # Faylni boshidan o‘qish uchun

        # 4. Modelga saqlash
        self.generated_document.docx_buyruq.save(f"{self.filename}_buyruq.docx", ContentFile(buffer.getvalue()),
                                                 save=True)

        # DOCXga QR-CODE joylashtirish va pdfga convert qilish
        pdf_filename = f"{self.filename}_buyruq"
        pdf_url = self.generate_qr_pdf(self.generated_document.docx_buyruq, pdf_filename)
        # Fayl yo‘lini nisbiy qilish (MEDIA_ROOT ni olib tashlash)
        relative_path = os.path.relpath(pdf_url, settings.MEDIA_ROOT).replace("\\", "/")
        # Faylni modelga saqlash
        with open(pdf_url, "rb") as f:
            self.generated_document.pdf_buyruq.save(relative_path, File(f), save=True)

        return self.generated_document

    # Dalolatnoma yaratish
    def gen_dalolatnoma(self):
        # 1. Shablonni ochish
        docx = DocxTemplate(self.dalolatnoma)

        # 2. Shablonni to'ldirish va saqlash
        self.context['qr_code'] = "{{qr_code}}"
        docx.render(self.context)

        # 3. Faylni xotirada saqlash
        buffer = BytesIO()
        docx.save(buffer)
        buffer.seek(0)  # Faylni boshidan o‘qish uchun

        # 4. Modelga saqlash
        self.generated_document.docx_dalolatnoma.save(f"{self.filename}_dalolatnoma.docx",
                                                      ContentFile(buffer.getvalue()), save=True)

        # DOCXga QR-CODE joylashtirish va pdfga convert qilish
        pdf_filename = f"{self.filename}_dalolatnoma"
        pdf_url = self.generate_qr_pdf(self.generated_document.docx_dalolatnoma, pdf_filename)
        # Fayl yo‘lini nisbiy qilish (MEDIA_ROOT ni olib tashlash)
        relative_path = os.path.relpath(pdf_url, settings.MEDIA_ROOT).replace("\\", "/")
        # Faylni modelga saqlash
        with open(pdf_url, "rb") as f:
            self.generated_document.pdf_dalolatnoma.save(relative_path, File(f), save=True)
        return self.generated_document

    # Bayonnoma yaratish
    def gen_bayonnoma(self):
        # 1. Shablonni ochish
        docx = DocxTemplate(self.bayonnoma)

        # 2. Shablonni to'ldirish va saqlash
        self.context['qr_code'] = "{{qr_code}}"
        docx.render(self.context)

        # 3. Faylni xotirada saqlash
        buffer = BytesIO()
        docx.save(buffer)
        buffer.seek(0)  # Faylni boshidan o‘qish uchun

        # 4. Modelga saqlash
        self.generated_document.docx_bayonnoma.save(f"{self.filename}_bayonnoma.docx",
                                                    ContentFile(buffer.getvalue()), save=True)

        # DOCXga QR-CODE joylashtirish va pdfga convert qilish
        pdf_filename = f"{self.filename}_bayonnoma"
        pdf_url = self.generate_qr_pdf(self.generated_document.docx_bayonnoma, pdf_filename)
        # Fayl yo‘lini nisbiy qilish (MEDIA_ROOT ni olib tashlash)
        relative_path = os.path.relpath(pdf_url, settings.MEDIA_ROOT).replace("\\", "/")
        # Faylni modelga saqlash
        with open(pdf_url, "rb") as f:
            self.generated_document.pdf_bayonnoma.save(relative_path, File(f), save=True)
        return self.generated_document


    # Gfafik yaratish
    def gen_grafik(self):
        # 1. Shablonni ochish
        docx = DocxTemplate(self.grafik)

        # 2. Shablonga EXCELdab olingan to'lov grafikni joylashtirish
        self.context['graphics'] = self.extract_excel()
        self.context['qr_code'] = "{{qr_code}}"
        docx.render(self.context)

        # 3. Faylni xotirada saqlash
        buffer = BytesIO()
        docx.save(buffer)
        buffer.seek(0)  # Faylni boshidan o‘qish uchun

        # 4. Modelga saqlash
        self.generated_document.docx_grafik.save(f"{self.filename}_grafik.docx", ContentFile(buffer.getvalue()),
                                                 save=True)
        # DOCXga QR-CODE joylashtirish va pdfga convert qilish
        pdf_filename = f"{self.filename}_grafik"
        pdf_url = self.generate_qr_pdf(self.generated_document.docx_grafik, pdf_filename)
        # Fayl yo‘lini nisbiy qilish (MEDIA_ROOT ni olib tashlash)
        relative_path = os.path.relpath(pdf_url, settings.MEDIA_ROOT).replace("\\", "/")
        # Faylni modelga saqlash
        with open(pdf_url, "rb") as f:
            self.generated_document.pdf_grafik.save(relative_path, File(f), save=True)

        return self.generated_document

    # EXCELdan ma'lumotlarni o'qish
    def extract_excel(self):
        xlsx = self.application.xlsx
        wb = load_workbook(filename=xlsx, read_only=True, data_only=True)
        ws = wb['График']

        # Cheklangan qatordan o'qish
        data = []

        for row in ws.iter_rows(min_row=9, max_row=100, min_col=1, max_col=6, values_only=True):
            obj = {
                "id": row[0],
                "date": row[1].strftime('%d.%m.%Y') if isinstance(row[1], datetime) else row[1],
                "total_payment": f"{row[2]:,.2f}" if isinstance(row[2], (int, float)) else row[2],
                "principal_balance": f"{row[3]:,.2f}" if isinstance(row[3], (int, float)) else row[3],
                "interest_payment": f"{row[4]:,.2f}" if isinstance(row[4], (int, float)) else row[4],
                "principal_payment": f"{row[5]:,.2f}" if isinstance(row[5], (int, float)) else row[5],
            }

            if obj['date'] != None and obj['id'] != None:
                data.append(obj)
            if obj["date"] == 'JAMI':
                obj['id'] = ""
                obj['total_payment'] = ""
                data.append(obj)
                break
        wb.close()
        print(data)
        return data

    def generate_qr_pdf(self, docx, pdf_filename):
        # QRCODE yaratish
        qr_file_name = f"{pdf_filename}.png"
        qr_file_path = os.path.join(settings.MEDIA_ROOT, "uploads/generated/qrcode", qr_file_name)

        # Papka mavjudligini tekshirish va yaratish
        os.makedirs(os.path.dirname(qr_file_path), exist_ok=True)

        # QR kod yaratish
        qr = qrcode.QRCode(
            version=4,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(f"http://{GLOBAL_IP}/media/uploads/generated/pdf/{pdf_filename}.pdf")
        qr.make(fit=True)

        # QR kodni tasvirga aylantirish va saqlash
        img = qr.make_image(fill="black", back_color="white")
        img.save(qr_file_path)

        # WORD templatega qrcode joylashtirish
        # 1. Shablonni ochish
        docx = DocxTemplate(docx)
        # 2. Tasvirni yuklash va o'lchamini belgilash
        image_path = qr_file_path
        image = InlineImage(docx, image_path, width=Mm(20), height=Mm(20))  # 50mm x 50mm
        self.context['qr_code'] = image
        docx.render(self.context)
        # Saqlash joyi
        docx_output_name = f"{self.filename}_upd.docx"
        docx_output_path = os.path.join(settings.MEDIA_ROOT, "uploads/generated/docx", docx_output_name)

        # Papka mavjudligini tekshirish va yaratish
        os.makedirs(os.path.dirname(docx_output_path), exist_ok=True)

        # Tayyorlangan hujjatni saqlash
        docx.save(docx_output_path)

        # Docxni pdfga aylantirish
        pdf_output_name = f"{pdf_filename}.pdf"
        pdf_output_path = os.path.join(settings.MEDIA_ROOT, pdf_output_name)

        pythoncom.CoInitialize()  # COM obyektlarini ishga tushirish
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False  # GUI'da ko'rinmasligi uchun

        convert(docx_output_path, pdf_output_path)

        return pdf_output_path

    def remove_temp_files(self):
        # Katalog yo'li
        time.sleep(2)
        directory = settings.MEDIA_ROOT
        for filename in os.listdir(directory):
            # Faylning to'liq yo'li
            file_path = os.path.join(directory, filename)

            # Fayl `.pdf` bilan tugasa va u haqiqiy fayl bo'lsa
            if filename.endswith('.pdf') and os.path.isfile(file_path):
                os.remove(file_path)  # Faylni o'chiradi
                print(f"{file_path} fayli o'chirildi.")
        return print('Barcha fayllar ochirildi.')

    def gen_excel_to_pdf(self):
        pdf_filename = f"{self.filename}_xulosa"
        xlsx = self.application.xlsx
        print(xlsx)
        wb = load_workbook(filename=xlsx, read_only=False, data_only=True)
        ws = wb['Заключение КК']
        # QR-kodni A1 katagiga joylashtirish

        # QRCODE yaratish
        qr_file_name = f"{pdf_filename}.png"
        qr_file_path = os.path.join(settings.MEDIA_ROOT, "uploads/generated/qrcode", qr_file_name)

        # Papka mavjudligini tekshirish va yaratish
        os.makedirs(os.path.dirname(qr_file_path), exist_ok=True)

        # QR kod yaratish
        qr = qrcode.QRCode(
            version=4,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(f"http://{GLOBAL_IP}/media/uploads/generated/pdf/{pdf_filename}.pdf")
        qr.make(fit=True)

        # QR kodni tasvirga aylantirish va saqlash
        img = qr.make_image(fill="black", back_color="white")
        img.save(qr_file_path)
        print(f"{qr_file_path}")

        img = Image(qr_file_path)
        # O'lchamlarni belgilash (rasmni o'lchamlari piksel bo'yicha)
        img.width = 100  # Kenglik
        img.height = 100  # Balandlik

        ws.add_image(img, "A53")


        # Excel faylni saqlash
        excel_file_name = f"{pdf_filename}.xlsx"
        excel_path = os.path.join(settings.MEDIA_ROOT, "uploads/generated/excel", excel_file_name)
        wb.save(excel_path)
        wb.close()  # openpyxl workbookni yopish

        pythoncom.CoInitialize()  # COM obyektlarini ishga tushirish

        # 1️⃣ Excel ilovasini ochish
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excel oynasi ko‘rinmasin

        # 2️⃣ Excel faylni ochish
        wb = excel.Workbooks.Open(excel_path)
        ws = wb.Sheets("Заключение КК")  # Faqat shu sheetni tanlash

        # 3️⃣ A4 formatga moslash va portret rejim
        ws.PageSetup.PaperSize = 9  # xlPaperA4
        ws.PageSetup.Orientation = 1  # xlPortrait (Portret)
        ws.PageSetup.Zoom = False  # Zoomni o‘chirib, Fit To ishlatish
        ws.PageSetup.FitToPagesWide = 1  # Kenglik bo‘yicha sahifaga sig‘dirish
        ws.PageSetup.FitToPagesTall = False  # Balans yo‘qotmaslik uchun

        # 4️⃣ Sahifa chekkalarini moslash
        ws.PageSetup.LeftMargin = excel.InchesToPoints(0.3)
        ws.PageSetup.RightMargin = excel.InchesToPoints(0.3)
        ws.PageSetup.TopMargin = excel.InchesToPoints(0.5)
        ws.PageSetup.BottomMargin = excel.InchesToPoints(0.5)

        # 5️⃣ Chop qilish chegaralarini belgilash
        ws.PageSetup.PrintArea = ws.UsedRange.Address  # Faqat ishlatilgan qismni chop etish

        gen_pdf_file = f"{pdf_filename}.pdf"
        pdf_path =  os.path.join(settings.MEDIA_ROOT, gen_pdf_file)
        ws.ExportAsFixedFormat(0, pdf_path)  # PDF formatiga saqlash (0 - xlTypePDF)

        # Excel'ni yopish
        wb.Close(SaveChanges=False)
        excel.Quit()
        del excel
        pythoncom.CoUninitialize()
        # Faylni modelga saqlash
        relative_path = os.path.relpath(pdf_path, settings.MEDIA_ROOT).replace("\\", "/")
        with open(pdf_path, "rb") as f:
            self.generated_document.pdf_xulosa.save(relative_path, File(f), save=True)

        return pdf_path

