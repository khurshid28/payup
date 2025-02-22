from datetime import datetime
from io import BytesIO
from pathlib import Path

import qrcode
from django.core.files.base import ContentFile
from docx.shared import Mm
from docxtpl import DocxTemplate, InlineImage
from docx2pdf import convert
from openpyxl.reader.excel import load_workbook

from payup.settings import BASE_DIR, MEDIA_ROOT, GLOBAL_IP


class GenDocument:
    def __init__(self, generated_document_doc_pdf=None, filename=None, shartnoma=None, buyruq=None, dalolatnoma=None,
                 grafik=None, context=None, report_obj=None):
        self.generated_document_doc_pdf = generated_document_doc_pdf
        self.filename = filename
        self.shartnoma = shartnoma
        self.buyruq = buyruq
        self.dalolatnoma = dalolatnoma
        self.grafik = grafik
        self.context = context
        self.report_obj = report_obj

    def display_info(self):
        return print(self.generated_document_doc_pdf)

    def gen_qrcode(self):
        # 1. Yangi GenDoc obyektini yaratamiz

        qr = qrcode.QRCode(
            version=4,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(f"{GLOBAL_IP}/contract/document_detail/{self.report_obj.unique_identifier}/")
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        buffer = BytesIO()
        img.save(buffer, format="PNG")

        # 4. QR kodni fayl sifatida saqlash
        qr_file = f"{self.filename}.png"
        self.generated_document_doc_pdf.filename = qr_file
        self.generated_document_doc_pdf.qrcode.save(qr_file, ContentFile(buffer.getvalue()),
                                                    save=True)  # Faylni modelga saqlaymiz
        return self.generated_document_doc_pdf

    # Shartnoma yaratish
    def gen_shartnoma(self):
        # 1. Shablonni ochish
        docx = DocxTemplate(self.shartnoma)

        # 2. Tasvirni yuklash va o'lchamini belgilash
        image_path = self.generated_document_doc_pdf.qrcode
        image = InlineImage(docx, image_path, width=Mm(30), height=Mm(30))  # 50mm x 50mm

        # Kontekst yaratish (tasvir o'zgaruvchisi bilan)
        self.context['qr_code'] = image

        # 3. Shablonni to'ldirish va saqlash
        docx.render(self.context)

        # 4. Faylni xotirada saqlash
        buffer = BytesIO()
        docx.save(buffer)
        buffer.seek(0)  # Faylni boshidan o‘qish uchun

        # 5. Modelga saqlash
        self.generated_document_doc_pdf.shartnoma_docx.save(f"{self.filename}_shartnoma.docx",
                                                            ContentFile(buffer.getvalue()), save=True)

        # 6. PDFga o'tkazish
        pdf_url = f"uploads/pdf/{self.filename}_shartnoma.pdf"
        convert(self.generated_document_doc_pdf.shartnoma_docx.path, MEDIA_ROOT / pdf_url)
        self.generated_document_doc_pdf.shartnoma_pdf = pdf_url
        self.generated_document_doc_pdf.save()

        return self.generated_document_doc_pdf

    # Buyruq yaratish
    def gen_buyruq(self):
        # 1. Shablonni ochish
        docx = DocxTemplate(self.buyruq)

        # 2. Tasvirni yuklash va o'lchamini belgilash
        image_path = self.generated_document_doc_pdf.qrcode
        image = InlineImage(docx, image_path, width=Mm(30), height=Mm(30))  # 50mm x 50mm

        # Kontekst yaratish (tasvir o'zgaruvchisi bilan)
        self.context['qr_code'] = image

        # 3. Shablonni to'ldirish va saqlash
        docx.render(self.context)

        # 4. Faylni xotirada saqlash
        buffer = BytesIO()
        docx.save(buffer)
        buffer.seek(0)  # Faylni boshidan o‘qish uchun

        # 5. Modelga saqlash
        self.generated_document_doc_pdf.buyruq_docx.save(f"{self.filename}_buyruq.docx", ContentFile(buffer.getvalue()),
                                                         save=True)

        # 6. PDFga o'tkazish
        pdf_url = f"uploads/pdf/{self.filename}_buyruq.pdf"
        convert(self.generated_document_doc_pdf.buyruq_docx.path, MEDIA_ROOT / pdf_url)
        self.generated_document_doc_pdf.buyruq_pdf = pdf_url
        self.generated_document_doc_pdf.save()
        return self.generated_document_doc_pdf

    # Dalolatnoma yaratish
    def gen_dalolatnoma(self):
        # 1. Shablonni ochish
        docx = DocxTemplate(self.dalolatnoma)

        # 2. Tasvirni yuklash va o'lchamini belgilash
        image_path = self.generated_document_doc_pdf.qrcode
        image = InlineImage(docx, image_path, width=Mm(30), height=Mm(30))  # 50mm x 50mm

        # Kontekst yaratish (tasvir o'zgaruvchisi bilan)
        self.context['qr_code'] = image

        # 3. Shablonni to'ldirish va saqlash
        docx.render(self.context)

        # 4. Faylni xotirada saqlash
        buffer = BytesIO()
        docx.save(buffer)
        buffer.seek(0)  # Faylni boshidan o‘qish uchun

        # 5. Modelga saqlash
        self.generated_document_doc_pdf.dalolatnoma_docx.save(f"{self.filename}_dalolatnoma.docx",
                                                              ContentFile(buffer.getvalue()), save=True)

        # 6. PDFga o'tkazish
        pdf_url = f"uploads/pdf/{self.filename}_dalolatnoma.pdf"
        convert(self.generated_document_doc_pdf.dalolatnoma_docx.path, MEDIA_ROOT / pdf_url)
        self.generated_document_doc_pdf.dalolatnoma_pdf = pdf_url
        self.generated_document_doc_pdf.save()

        return self.generated_document_doc_pdf

    # Gfafik yaratish
    def gen_grafik(self):
        # 1. Shablonni ochish
        docx = DocxTemplate(self.grafik)

        # 2. Tasvirni yuklash va o'lchamini belgilash
        image_path = self.generated_document_doc_pdf.qrcode
        image = InlineImage(docx, image_path, width=Mm(30), height=Mm(30))  # 50mm x 50mm

        # Kontekst yaratish (tasvir o'zgaruvchisi bilan)
        self.context['qr_code'] = image

        # Shablonga to'lov grafikni joylashtirish
        self.context['graphics'] = self.extract_excel()

        # 3. Shablonni to'ldirish va saqlash
        docx.render(self.context)

        # 4. Faylni xotirada saqlash
        buffer = BytesIO()
        docx.save(buffer)
        buffer.seek(0)  # Faylni boshidan o‘qish uchun

        # 5. Modelga saqlash
        self.generated_document_doc_pdf.grafik_docx.save(f"{self.filename}_grafik.docx", ContentFile(buffer.getvalue()),
                                                         save=True)

        # 6. PDFga o'tkazish
        pdf_url = f"uploads/pdf/{self.filename}_grafik.pdf"
        convert(self.generated_document_doc_pdf.grafik_docx.path, MEDIA_ROOT / pdf_url)
        self.generated_document_doc_pdf.grafik_pdf = pdf_url
        self.generated_document_doc_pdf.save()
        return self.generated_document_doc_pdf.id

    # EXCELdan ma'lumotlarni o'qish
    def extract_excel(self):
        xlsx = self.report_obj.xlsx
        wb = load_workbook(filename=xlsx, read_only=True, data_only=True)
        ws = wb['График']

        # Cheklangan qatordan o'qish
        data = []

        for row in ws.iter_rows(min_row=9, max_row=100, min_col=1, max_col=6, values_only=True):
            obj = {
                "id": row[0],
                "date": row[1].strftime('%d.%m.%Y') if type(row[1]) is datetime else row[1],
                "total_payment": '{:,.2f}'.format(row[2]) if type(row[2]) is float else row[2],
                "principal_balance": '{:,.2f}'.format(row[3]) if isinstance(row[3], float) else '{:,.0f}'.format(row[3]),
                "interest_payment": '{:,.2f}'.format(row[4]) if isinstance(row[4], float) else '{:,.0f}'.format(row[4]),
                "principal_payment": '{:,.2f}'.format(row[5]) if isinstance(row[5], float) else '{:,.0f}'.format(row[5]),
            }
            if obj['date'] != None and obj['id'] != None:
                data.append(obj)
            if obj["date"] == 'JAMI':
                obj['id'] = ""
                obj['total_payment'] = ""
                data.append(obj)
                break
        wb.close()
        print(data)
        return data
